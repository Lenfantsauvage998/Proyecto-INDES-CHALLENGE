"""
FastAPI backend for ISSE Certificate Tool.

Endpoints:
  GET  /api/stats          - DB summary counts
  GET  /api/ciclos         - distinct ciclos list
  GET  /api/professors     - distinct professor names
  GET  /api/certificates   - query with ?profesor=&ciclo=
  POST /api/upload         - upload new Excel, run ETL in append mode
  GET  /api/export         - export query results as xlsx or csv
  GET  /api/debug          - step-by-step ETL trace for a professor+ciclo
"""

import io
import os
import sys
import tempfile
import threading
import uuid
from pathlib import Path
from dotenv import load_dotenv

load_dotenv()

import pandas as pd
import psycopg2
import psycopg2.extras
from fastapi import FastAPI, File, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse

# ── config ────────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent.parent
DATABASE_URL = os.environ.get("DATABASE_URL", "")
sys.path.insert(0, str(BASE_DIR))

from etl import run_etl_on_file, parse_hour  # noqa: E402

# ── app ───────────────────────────────────────────────────────────────────────
app = FastAPI(title="ISSE Certificate API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── background upload job tracker ────────────────────────────────────────────
_upload_jobs: dict[str, dict] = {}


def get_conn():
    if not DATABASE_URL:
        raise HTTPException(503, "DATABASE_URL no configurada.")
    conn = psycopg2.connect(DATABASE_URL)
    return conn


def _fetchall(conn, sql: str, params=None) -> list:
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(sql, params or [])
    return cur.fetchall()


def _fetchone_tuple(conn, sql: str, params=None):
    cur = conn.cursor()
    cur.execute(sql, params or [])
    return cur.fetchone()


def _profesor_filter(q: str) -> tuple[str, list]:
    """Return (sql_fragment, params) matching all tokens regardless of order."""
    tokens = [t for t in q.lower().split() if t]
    if not tokens:
        return "", []
    clauses = " AND ".join("LOWER(profesor) LIKE %s" for _ in tokens)
    params = [f"%{t}%" for t in tokens]
    return f" AND ({clauses})", params


# ── endpoints ────────────────────────────────────────────────────────────────
@app.get("/api/stats")
def stats():
    try:
        conn = get_conn()
        row = _fetchone_tuple(
            conn,
            "SELECT COUNT(*) AS total, COUNT(DISTINCT profesor) AS profs, "
            "COUNT(DISTINCT ciclo_lectivo) AS ciclos FROM certificados",
        )
        conn.close()
        return {
            "total_records": row[0],
            "total_professors": row[1],
            "total_ciclos": row[2],
        }
    except HTTPException:
        raise
    except Exception as exc:
        return {"total_records": 0, "total_professors": 0, "total_ciclos": 0, "error": str(exc)}


@app.get("/api/ciclos")
def ciclos():
    if not DATABASE_URL:
        return {"ciclos": []}
    try:
        conn = psycopg2.connect(DATABASE_URL)
        cur = conn.cursor()
        cur.execute("SELECT DISTINCT ciclo_lectivo FROM certificados ORDER BY ciclo_lectivo")
        rows = cur.fetchall()
        conn.close()
    except Exception:
        rows = []
    return {"ciclos": [r[0] for r in rows]}


@app.get("/api/professors")
def professors(
    q: str = Query("", description="Partial name filter"),
    ciclo: str = Query("", description="Single ciclo filter (legacy)"),
    ciclos: str = Query("", description="Comma-separated exact ciclo list"),
):
    ciclo_list = [c.strip() for c in ciclos.split(",") if c.strip()] if ciclos else []
    if ciclo.strip() and ciclo.strip() not in ciclo_list:
        ciclo_list.append(ciclo.strip())

    conn = get_conn()
    sql = "SELECT DISTINCT profesor FROM certificados WHERE 1=1"
    params: list = []
    if q:
        frag, fparams = _profesor_filter(q.strip())
        sql += frag
        params.extend(fparams)
    if ciclo_list:
        placeholders = ",".join(["%s"] * len(ciclo_list))
        sql += f" AND ciclo_lectivo IN ({placeholders})"
        params.extend(ciclo_list)
    sql += " ORDER BY profesor LIMIT 50"
    rows = _fetchall(conn, sql, params)
    conn.close()
    return {"professors": [r["profesor"] for r in rows if r["profesor"]]}


@app.get("/api/certificates")
def certificates(
    profesor: str = Query("", description="Professor name partial match"),
    ciclo: str = Query("", description="Single ciclo (legacy, partial match)"),
    ciclos: str = Query("", description="Comma-separated exact ciclo list"),
):
    ciclo_list = [c.strip() for c in ciclos.split(",") if c.strip()] if ciclos else []
    if ciclo.strip() and ciclo.strip() not in ciclo_list:
        ciclo_list.append(ciclo.strip())

    if not profesor.strip() and not ciclo_list:
        raise HTTPException(400, "Provide at least one filter: profesor or ciclo.")

    conn = get_conn()
    # Check if nombre_curso column exists (safe, uses information_schema)
    row_check = _fetchone_tuple(
        conn,
        "SELECT COUNT(*) FROM information_schema.columns "
        "WHERE table_name='certificados' AND column_name='nombre_curso'",
    )
    nombre_col = (
        "COALESCE(nombre_curso, materia) AS nombre_curso"
        if row_check and row_check[0] > 0
        else "materia AS nombre_curso"
    )

    sql = (
        f"SELECT profesor, componente, materia, {nombre_col}, ciclo_lectivo, "
        "horas_semestre, num_grupos, departamento, fecha_inicio, fecha_fin "
        "FROM certificados WHERE 1=1"
    )
    params: list = []
    if profesor.strip():
        frag, fparams = _profesor_filter(profesor.strip())
        sql += frag
        params.extend(fparams)
    if ciclo_list:
        placeholders = ",".join(["%s"] * len(ciclo_list))
        sql += f" AND ciclo_lectivo IN ({placeholders})"
        params.extend(ciclo_list)
    sql += " ORDER BY profesor, ciclo_lectivo, nombre_curso"
    rows = _fetchall(conn, sql, params)
    conn.close()
    return {"results": [dict(r) for r in rows]}


@app.get("/api/bot/certificate")
def bot_certificate(profesor: str = Query(..., description="Professor name partial match")):
    from pdf_generator import generate_pdf_bytes
    from fastapi.responses import Response
    
    if not profesor.strip():
        raise HTTPException(400, "Provide a profesor name.")
        
    # Reuse the logic from the certificates endpoint
    try:
        cert_data = certificates(profesor=profesor, ciclo="", ciclos="")
        records = cert_data.get("results", [])
    except HTTPException as e:
        raise e
        
    if not records:
        raise HTTPException(404, "No records found for the given professor.")
        
    # Get the actual professor name from the first record
    real_profesor_name = records[0]["profesor"]
    
    pdf_bytes = generate_pdf_bytes(real_profesor_name, records)
    
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="certificado_{real_profesor_name.replace(" ", "_")}.pdf"'}
    )


@app.post("/api/upload")
async def upload(file: UploadFile = File(...)):
    if not file.filename or not (
        file.filename.endswith(".xlsx") or file.filename.endswith(".xls")
    ):
        raise HTTPException(400, "Solo se aceptan archivos .xlsx o .xls")

    # ── stream to disk in chunks — zero synchronous pandas work here ──────────
    tmp_fd, tmp_name = tempfile.mkstemp(suffix=".xlsx")
    tmp_path = Path(tmp_name)
    try:
        with os.fdopen(tmp_fd, "wb") as f:
            while True:
                chunk = await file.read(512 * 1024)  # 512KB chunks
                if not chunk:
                    break
                f.write(chunk)
    except Exception as exc:
        tmp_path.unlink(missing_ok=True)
        raise HTTPException(400, f"Error recibiendo el archivo: {exc}") from exc

    # ── launch background job immediately ─────────────────────────────────────
    job_id = str(uuid.uuid4())[:8]
    _upload_jobs[job_id] = {"status": "processing", "result": None, "error": None}

    def _run_etl(path: Path, jid: str):
        try:
            # Validate: check Ciclo Lectivo column
            preview_df = pd.read_excel(path, engine="calamine", usecols=["Ciclo Lectivo"])
            incoming = set(str(c) for c in preview_df["Ciclo Lectivo"].dropna().unique())
            del preview_df

            if not incoming:
                _upload_jobs[jid] = {
                    "status": "error",
                    "result": None,
                    "error": "El archivo no contiene columna 'Ciclo Lectivo' o está vacío.",
                }
                return

            # Check conflicts against PostgreSQL
            try:
                _conn = psycopg2.connect(DATABASE_URL)
                _cur = _conn.cursor()
                _cur.execute("SELECT DISTINCT ciclo_lectivo FROM certificados")
                existing = {r[0] for r in _cur.fetchall()}
                _conn.close()
            except Exception:
                existing = set()

            overlap = sorted(incoming & existing)
            if overlap:
                _upload_jobs[jid] = {
                    "status": "conflict",
                    "result": None,
                    "error": "El archivo contiene periodos ya cargados en la base de datos.",
                    "conflicting_ciclos": overlap,
                }
                return

            # Run ETL (db_path ignored — uses DATABASE_URL internally)
            result = run_etl_on_file(path, append=True)
            _upload_jobs[jid] = {"status": "done", "result": result, "error": None}
        except Exception as exc:
            _upload_jobs[jid] = {"status": "error", "result": None, "error": str(exc)}
        finally:
            path.unlink(missing_ok=True)

    threading.Thread(target=_run_etl, args=(tmp_path, job_id), daemon=True).start()
    return {"job_id": job_id, "status": "processing"}


@app.get("/api/upload-status/{job_id}")
def upload_status(job_id: str):
    job = _upload_jobs.get(job_id)
    if not job:
        raise HTTPException(404, "Job no encontrado.")
    return job


@app.delete("/api/ciclos/{ciclo_lectivo:path}")
def delete_ciclo(ciclo_lectivo: str):
    """Delete all records for a specific ciclo_lectivo."""
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM certificados WHERE ciclo_lectivo = %s", (ciclo_lectivo,))
        deleted = cur.rowcount
        # Also clean raw_rows if table exists
        cur.execute(
            "SELECT 1 FROM information_schema.tables WHERE table_name = 'raw_rows'"
        )
        if cur.fetchone():
            cur.execute("DELETE FROM raw_rows WHERE ciclo_lectivo = %s", (ciclo_lectivo,))
        conn.commit()
    except Exception as exc:
        conn.rollback()
        conn.close()
        raise HTTPException(500, str(exc)) from exc
    conn.close()
    if deleted == 0:
        raise HTTPException(404, f"Periodo '{ciclo_lectivo}' no encontrado en la base de datos.")
    return {"message": f"Periodo '{ciclo_lectivo}' eliminado.", "deleted_records": deleted}


@app.delete("/api/db")
def reset_db():
    """Wipe all data from certificados and raw_rows tables."""
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("DELETE FROM certificados")
        # Delete raw_rows only if table exists
        cur.execute(
            "SELECT 1 FROM information_schema.tables WHERE table_name = 'raw_rows'"
        )
        if cur.fetchone():
            cur.execute("DELETE FROM raw_rows")
        conn.commit()
        conn.close()
    except Exception as exc:
        raise HTTPException(500, f"No se pudo limpiar la base de datos: {exc}") from exc
    return {"message": "Base de datos vaciada correctamente."}


def _build_certificate_excel(df: "pd.DataFrame", buf: "io.BytesIO") -> None:
    """Build Excel certificate: professor header + 7-column data table."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    NUM_COLS = 7  # SEMESTRE COMPONENTE ASIGNATURA SESIONES DEPARTAMENTO FECHA_INICIO FECHA_FIN

    wb = Workbook()
    ws = wb.active
    ws.title = "Certificado"

    # ── styles ────────────────────────────────────────────────────────────────
    title_font  = Font(name="Arial", bold=True, size=13)
    info_font   = Font(name="Arial", bold=True, size=11)
    hdr_font    = Font(name="Arial", bold=True, size=11)
    data_font   = Font(name="Arial", size=10)
    hdr_fill    = PatternFill("solid", fgColor="BFBFBF")
    thin        = Side(style="thin",   color="000000")
    medium      = Side(style="medium", color="000000")
    border      = Border(left=thin,   right=thin,   top=thin,   bottom=thin)
    hdr_border  = Border(left=medium, right=medium, top=medium, bottom=medium)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── professor info (from first row) ───────────────────────────────────────
    profesor = str(df["profesor"].iloc[0]) if "profesor" in df.columns and len(df) else ""
    id_prof  = str(df["num_doc_docente"].iloc[0]) if "num_doc_docente" in df.columns and len(df) else ""
    try:
        id_prof = str(int(float(id_prof))) if id_prof not in ("", "nan", "None") else "—"
    except Exception:
        id_prof = id_prof or "—"

    # ── ROW 1: Title ──────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
    c = ws.cell(row=1, column=1, value="PROCESO DE CERTIFICACIÓN · FACULTAD DE INGENIERÍA")
    c.font = title_font
    c.alignment = center
    ws.row_dimensions[1].height = 30

    # ── ROW 2: Docente name (cols 1-4) · Nº Documento (cols 5-7) ─────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    c2a = ws.cell(row=2, column=1, value=f"Docente:  {profesor}")
    c2a.font = info_font
    c2a.alignment = left

    ws.merge_cells(start_row=2, start_column=5, end_row=2, end_column=NUM_COLS)
    c2b = ws.cell(row=2, column=5, value=f"Nº Documento:  {id_prof}")
    c2b.font = info_font
    c2b.alignment = left
    ws.row_dimensions[2].height = 20

    # ── ROW 3: visual gap ─────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 6

    # ── ROW 4: Column headers ─────────────────────────────────────────────────
    headers = ["SEMESTRE", "COMPONENTE", "ASIGNATURA", "SESIONES",
               "DEPARTAMENTO", "FECHA INICIO", "FECHA FIN"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
        c.border = hdr_border
    ws.row_dimensions[4].height = 24

    def fmt_date(d: object) -> str:
        s = str(d) if d is not None else ""
        if s in ("", "None", "nan"):
            return ""
        try:
            y, m, day = s[:10].split("-")
            return f"{day}/{m}/{y}"
        except Exception:
            return s

    # ── DATA ROWS (sorted ciclo low → high, merged SEMESTRE column) ───────────
    row_num = 5
    for ciclo, group in df.groupby("ciclo_lectivo", sort=True):
        start_row = row_num
        for _, record in group.iterrows():
            vals = [
                str(ciclo),
                (record.get("componente") or "").upper(),
                (record.get("nombre_curso") or record.get("materia") or "").upper(),
                int(record["horas_semestre"]),
                (record.get("departamento") or "").upper(),
                fmt_date(record.get("fecha_inicio")),
                fmt_date(record.get("fecha_fin")),
            ]
            aligns = [center, center, left, center, left, center, center]

            for col, (val, aln) in enumerate(zip(vals, aligns), 1):
                c = ws.cell(row=row_num, column=col, value=val)
                c.font = data_font
                c.alignment = aln
                c.border = border

            row_num += 1

        if row_num - start_row > 1:
            ws.merge_cells(
                start_row=start_row, start_column=1,
                end_row=row_num - 1, end_column=1,
            )
            ws.cell(row=start_row, column=1).alignment = center

    # ── column widths ─────────────────────────────────────────────────────────
    col_widths = [20, 16, 36, 11, 36, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(buf)


@app.get("/api/export")
def export(
    profesor: str = Query(""),
    ciclo: str = Query("", description="Single ciclo (legacy)"),
    ciclos: str = Query("", description="Comma-separated exact ciclo list"),
    format: str = Query("excel", pattern="^(excel|csv)$"),
):
    ciclo_list = [c.strip() for c in ciclos.split(",") if c.strip()] if ciclos else []
    if ciclo.strip() and ciclo.strip() not in ciclo_list:
        ciclo_list.append(ciclo.strip())

    if not profesor.strip() and not ciclo_list:
        raise HTTPException(400, "Provide at least one filter.")

    conn = get_conn()
    sql = (
        "SELECT profesor, id_profesor, num_doc_docente, componente, materia, "
        "COALESCE(nombre_curso, materia) AS nombre_curso, "
        "ciclo_lectivo, horas_semestre, num_grupos, departamento, fecha_inicio, fecha_fin "
        "FROM certificados WHERE 1=1"
    )
    params: list = []
    if profesor.strip():
        frag, fparams = _profesor_filter(profesor.strip())
        sql += frag
        params.extend(fparams)
    if ciclo_list:
        placeholders = ",".join(["%s"] * len(ciclo_list))
        sql += f" AND ciclo_lectivo IN ({placeholders})"
        params.extend(ciclo_list)
    sql += " ORDER BY ciclo_lectivo, nombre_curso"
    df = pd.read_sql_query(sql, conn, params=params)
    conn.close()

    buf = io.BytesIO()
    if format == "excel":
        _build_certificate_excel(df, buf)
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = "certificado.xlsx"
    else:
        csv_cols = ["profesor", "num_doc_docente", "ciclo_lectivo", "componente",
                    "nombre_curso", "horas_semestre", "departamento", "fecha_inicio", "fecha_fin"]
        out = df[[c for c in csv_cols if c in df.columns]].copy()
        out.columns = ["DOCENTE", "Nº DOCUMENTO", "SEMESTRE", "COMPONENTE",
                       "ASIGNATURA", "SESIONES", "DEPARTAMENTO",
                       "FECHA INICIO", "FECHA FIN"][:len(out.columns)]
        out.to_csv(buf, index=False, encoding="utf-8-sig")
        media = "text/csv"
        filename = "certificado.csv"

    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type=media,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── /api/debug ────────────────────────────────────────────────────────────────
@app.get("/api/debug")
def debug_etl(
    profesor: str = Query(..., description="Partial professor name"),
    ciclo: str = Query(..., description="Ciclo lectivo e.g. PERIODO 2025-1"),
):
    import traceback, json
    try:
        result = _debug_etl_inner(profesor, ciclo)
        return JSONResponse(content=json.loads(json.dumps(result, default=str)))
    except Exception as exc:
        raise HTTPException(500, detail=f"{type(exc).__name__}: {exc}\n\n{traceback.format_exc()}")


def _debug_etl_inner(profesor: str, ciclo: str):
    import gc
    import numpy as np

    conn = get_conn()
    cur = conn.cursor()

    # Check raw_rows table exists
    cur.execute(
        "SELECT 1 FROM information_schema.tables WHERE table_name = 'raw_rows'"
    )
    if not cur.fetchone():
        conn.close()
        raise HTTPException(
            503,
            "Sube un archivo Excel primero para activar esta función.",
        )

    # Load filtered raw rows from PostgreSQL
    df_raw_sql = pd.read_sql_query(
        "SELECT * FROM raw_rows "
        "WHERE ciclo_lectivo LIKE %s AND UPPER(nombre_profesor) LIKE %s",
        conn,
        params=(f"%{ciclo}%", f"%{profesor.upper().strip()}%"),
    )

    # Fetch global dup hashes (hashes that appear more than once across all rows)
    cur2 = conn.cursor()
    cur2.execute(
        "SELECT row_hash FROM raw_rows GROUP BY row_hash HAVING COUNT(*) > 1"
    )
    dup_hashes = {r[0] for r in cur2.fetchall()}
    conn.close()

    # Rename DB column names back to original Excel column names
    RENAME_BACK = {
        "nombre_profesor":    "Nombre profesor",
        "ciclo_lectivo_raw":  "Ciclo Lectivo",
        "num_clase":          "Nº Clase",
        "id_seccion_combinada": "ID Sección Combinada",
        "hora_inicio":        "Hora Inicio",
        "hora_final":         "Hora Final",
        "clase_principal":    "Clase Principal",
        "descripcion_materia":"Descripción Materia",
        "nombre_curso":       "Nombre del curso",
        "dia":                "Día",
        "id_instalacion":     "ID Instalación",
        "componente_desc":    "Componente Descripción",
        "tipo_dedicacion":    "Tipo dedicación",
        "id_profesor":        "Id profesor",
        "descripcion_1":      "Descripción.1",
    }
    hashes = df_raw_sql["row_hash"].values
    df_raw = df_raw_sql.rename(columns=RENAME_BACK).drop(
        columns=["ciclo_lectivo", "row_hash"], errors="ignore"
    )
    del df_raw_sql; gc.collect()

    # ── helpers ───────────────────────────────────────────────────────────────
    def _safe(v):
        if v is None:
            return None
        try:
            if pd.isna(v):
                return None
        except (TypeError, ValueError):
            pass
        if isinstance(v, np.integer):
            return int(v)
        if isinstance(v, np.floating):
            f = float(v)
            return None if (np.isnan(f) or np.isinf(f)) else round(f, 4)
        if isinstance(v, float):
            return None if (np.isnan(v) or np.isinf(v)) else round(v, 4)
        if isinstance(v, np.bool_):
            return bool(v)
        if hasattr(v, "isoformat"):
            return str(v)
        return v

    def to_records(df, cols):
        available = [c for c in cols if c in df.columns]
        raw = df[available].to_dict("records")
        return [
            {k: _safe(v) for k, v in row.items()}
            for row in raw
        ], available

    def find_col_nonempty(df, *candidates):
        """Return first candidate that exists in df AND has at least one non-null value."""
        for c in candidates:
            if c in df.columns and df[c].notna().any():
                return c
        return None

    BASE_COLS = [
        "Nombre profesor", "Ciclo Lectivo", "Nº Clase",
        "ID Sección Combinada", "Hora Inicio", "Hora Final",
        "Clase Principal", "Descripción Materia", "Nombre del curso",
    ]

    # Detect optional columns
    day_col  = find_col_nonempty(df_raw, "Día", "Dia", "Día de la semana", "Dia semana")
    inst_col = find_col_nonempty(df_raw, "ID Instalación", "ID Instalacion", "ID Instalacion ", "Instalación")

    display_cols = BASE_COLS.copy()
    if day_col:
        display_cols.insert(4, day_col)
    if inst_col:
        display_cols.append(inst_col)

    steps = []

    # ══ STEP 1 — Raw rows ════════════════════════════════════════════════════
    rows, cols = to_records(df_raw, display_cols)
    steps.append({
        "step": 1,
        "label": "Datos crudos",
        "description": f"Todas las filas para el profesor '{profesor}' en el ciclo '{ciclo}' sin ninguna transformación.",
        "row_count": len(df_raw),
        "removed_count": 0,
        "removed_rows": [],
        "highlight_cols": [],
        "columns": cols,
        "rows": rows,
    })

    # ══ STEP 2 — Global dedup (reconstructed via row_hash) ════════════════════
    dup_series  = pd.Series(hashes)
    is_dup_hash = dup_series.isin(dup_hashes)
    is_dup_pos  = dup_series.duplicated(keep="first")
    removed     = is_dup_hash & is_dup_pos
    df_dedup      = df_raw[~removed.values].copy()
    dropped_dedup = df_raw[removed.values].copy()
    del df_raw; gc.collect()

    dr, _ = to_records(dropped_dedup, display_cols)
    rows, cols = to_records(df_dedup, display_cols)
    steps.append({
        "step": 2,
        "label": "Eliminar duplicados exactos",
        "description": "Se eliminan filas 100% idénticas en todas sus columnas. Opera sobre el dataset completo antes de separar por semestre.",
        "row_count": len(df_dedup),
        "removed_count": len(dropped_dedup),
        "removed_rows": dr,
        "highlight_cols": [],
        "columns": cols,
        "rows": rows,
    })

    # ══ STEP 1b — Semester isolation (visual checkpoint) ═════════════════════
    rows, cols = to_records(df_dedup, display_cols)
    steps.append({
        "step": 3,
        "label": "Aislamiento por semestre",
        "description": f"El dataset completo se divide por Ciclo Lectivo. Solo se trabaja con '{ciclo}'. Esto impide que un Nº Clase igual en otro semestre colisione con este.",
        "row_count": len(df_dedup),
        "removed_count": 0,
        "removed_rows": [],
        "highlight_cols": ["Ciclo Lectivo"],
        "columns": cols,
        "rows": rows,
    })

    # ══ STEP 4 — Assign grupo_id ══════════════════════════════════════════════
    sem_df = df_dedup.copy()
    has_combined = (
        sem_df["ID Sección Combinada"].notna()
        & (sem_df["ID Sección Combinada"].astype(str).str.strip() != "")
    )
    sem_df["grupo_id"] = sem_df["ID Sección Combinada"].where(
        has_combined, sem_df["Nº Clase"].astype(str)
    )
    cols_g = display_cols + ["grupo_id"]
    rows, cols = to_records(sem_df, cols_g)
    steps.append({
        "step": 4,
        "label": "Asignar grupo_id",
        "description": "Si 'ID Sección Combinada' tiene valor → grupo_id = ese valor. Si está vacío/nulo → grupo_id = Nº Clase. Este campo es la clave de agrupación.",
        "row_count": len(sem_df),
        "removed_count": 0,
        "removed_rows": [],
        "highlight_cols": ["grupo_id", "ID Sección Combinada"],
        "columns": cols,
        "rows": rows,
    })

    # ══ STEP 5 — Dedup combined sections ════════════════════════════════════
    if inst_col and inst_col in sem_df.columns:
        dedup_subset = ["grupo_id", "Hora Inicio", "Hora Final", inst_col]
        combined_before = sem_df[has_combined].copy()
        combined_after  = combined_before.drop_duplicates(subset=dedup_subset)
        plain            = sem_df[~has_combined].copy()
        dropped_combined = combined_before[~combined_before.index.isin(combined_after.index)]
        sem_df2 = pd.concat([combined_after, plain], ignore_index=True)
        dr, _ = to_records(dropped_combined, cols_g)
        rows, cols = to_records(sem_df2, cols_g)
        steps.append({
            "step": 5,
            "label": "Dedup secciones combinadas",
            "description": f"Dentro de cada ID Sección Combinada: si dos filas tienen igual grupo_id + Hora Inicio + Hora Final + {inst_col}, son la MISMA sesión física repetida para cursos fusionados. Se elimina el duplicado — no se suman las horas.",
            "row_count": len(sem_df2),
            "removed_count": len(dropped_combined),
            "removed_rows": dr,
            "highlight_cols": ["grupo_id", "Hora Inicio", "Hora Final", inst_col],
            "columns": cols,
            "rows": rows,
        })
    else:
        sem_df2 = sem_df.copy()
        rows, cols = to_records(sem_df2, cols_g)
        steps.append({
            "step": 5,
            "label": "Dedup secciones combinadas",
            "description": f"Columna de instalación no encontrada ({inst_col}). Paso omitido — verifica el nombre exacto de la columna en el Excel.",
            "row_count": len(sem_df2),
            "removed_count": 0,
            "removed_rows": [],
            "highlight_cols": [],
            "columns": cols,
            "rows": rows,
        })

    # ══ STEP 6 — Parse times + duration ═════════════════════════════════════
    sem_df2["_h_ini"] = sem_df2["Hora Inicio"].apply(parse_hour)
    sem_df2["_h_fin"] = sem_df2["Hora Final"].apply(parse_hour)
    sem_df2["_dur"]   = (sem_df2["_h_fin"] - sem_df2["_h_ini"]).clip(lower=0)
    time_cols = cols_g + ["_h_ini", "_h_fin", "_dur"]
    rows, cols = to_records(sem_df2, time_cols)
    steps.append({
        "step": 6,
        "label": "Parsear horas + duración",
        "description": "Hora Inicio / Hora Final → decimal militar (e.g. '01:00:PM' → 13.0). _dur = Hora Final − Hora Inicio, mínimo 0.",
        "row_count": len(sem_df2),
        "removed_count": 0,
        "removed_rows": [],
        "highlight_cols": ["_h_ini", "_h_fin", "_dur"],
        "columns": cols,
        "rows": rows,
    })

    # ══ STEP 7 — Aggregate + × 16 ═════════════════════════════════════════════
    if "Descripción.1" in sem_df2.columns:
        sem_df2["_dpto"] = sem_df2["Descripción.1"].fillna("").astype(str)
    elif "Departamento" in sem_df2.columns:
        sem_df2["_dpto"] = sem_df2["Departamento"].astype(str)
    else:
        sem_df2["_dpto"] = ""

    if "Nombre del curso" in sem_df2.columns:
        sem_df2 = sem_df2.rename(columns={"Nombre del curso": "nombre_curso"})

    cert_keys = [k for k in [
        "Ciclo Lectivo", "grupo_id", "Nombre profesor", "Id profesor",
        "Descripción Materia", "nombre_curso", "Componente Descripción",
        "_dpto", "Tipo dedicación",
    ] if k in sem_df2.columns]

    cert = (
        sem_df2.groupby(cert_keys, dropna=False)
        .agg(horas_semana=("_dur", "sum"), num_grupos=("Nº Clase", "nunique"))
        .reset_index()
    )
    cert["horas_semestre"] = cert["horas_semana"] * 16

    agg_show = ["grupo_id", "Descripción Materia", "nombre_curso",
                "horas_semana", "horas_semestre", "num_grupos"]
    rows, cols = to_records(cert, agg_show)
    steps.append({
        "step": 7,
        "label": "Agregación final × 16",
        "description": "Se agrupa por (ciclo, grupo_id, profesor, …). Las duraciones se suman → horas_semana (horas semanales). Luego horas_semana × 16 = horas_semestre.",
        "row_count": len(cert),
        "removed_count": 0,
        "removed_rows": [],
        "highlight_cols": ["horas_semana", "horas_semestre"],
        "columns": cols,
        "rows": rows,
    })

    # ══ STEP 8 — Merge same materia+nombre_curso ══════════════════════════════
    cert_renamed = cert.rename(columns={
        "Nombre profesor":      "profesor",
        "Id profesor":          "id_profesor",
        "Ciclo Lectivo":        "ciclo_lectivo",
        "Descripción Materia":  "materia",
        "Componente Descripción": "componente",
        "_dpto":                "departamento",
        "Tipo dedicación":      "tipo_dedicacion",
    })

    merge_keys = [k for k in [
        "ciclo_lectivo", "profesor", "id_profesor", "materia", "nombre_curso",
        "componente", "departamento", "tipo_dedicacion",
    ] if k in cert_renamed.columns]

    if merge_keys:
        cert_merged = (
            cert_renamed.groupby(merge_keys, dropna=False)
            .agg(horas_semestre=("horas_semestre", "sum"), num_grupos=("num_grupos", "sum"))
            .reset_index()
        )

        merge_id_cols = [k for k in ["materia", "nombre_curso"] if k in cert_renamed.columns]
        if merge_id_cols:
            dup_mask = cert_renamed.duplicated(subset=merge_id_cols, keep=False)
            merged_source_rows = cert_renamed[dup_mask]
        else:
            merged_source_rows = cert_renamed.iloc[0:0]

        dr8, _   = to_records(merged_source_rows, ["materia", "nombre_curso", "horas_semestre", "num_grupos"])
        rows8, cols8 = to_records(cert_merged,       ["materia", "nombre_curso", "horas_semestre", "num_grupos"])
        removed8 = len(cert_renamed) - len(cert_merged)
        steps.append({
            "step": 8,
            "label": "Fusionar cursos duplicados (mismo nombre + materia)",
            "description": "Si quedan dos filas con la misma Descripción Materia Y nombre_curso (p.ej. el profesor tenía dos secciones del mismo curso), se suman sus horas_semestre. El resultado es UN solo renglón de certificado por curso.",
            "row_count": len(cert_merged),
            "removed_count": removed8,
            "removed_rows": dr8,
            "highlight_cols": ["materia", "nombre_curso", "horas_semestre"],
            "columns": cols8,
            "rows": rows8,
        })
    else:
        rows8, cols8 = to_records(cert_renamed, ["materia", "nombre_curso", "horas_semestre"])
        steps.append({
            "step": 8,
            "label": "Fusionar cursos duplicados (mismo nombre + materia)",
            "description": "Paso omitido — columnas de agrupación no disponibles.",
            "row_count": len(cert_renamed),
            "removed_count": 0,
            "removed_rows": [],
            "highlight_cols": [],
            "columns": cols8,
            "rows": rows8,
        })

    return {
        "profesor": profesor,
        "ciclo": ciclo,
        "excel": "PostgreSQL (raw_rows)",
        "day_col": day_col,
        "inst_col": inst_col,
        "steps": steps,
    }


# ── static files (React build) — must be LAST so /api/* routes take priority ──
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse as _FileResponse

_STATIC_DIR = BASE_DIR / "frontend" / "dist"

if _STATIC_DIR.exists():
    app.mount("/assets", StaticFiles(directory=_STATIC_DIR / "assets"), name="assets")

    @app.get("/{full_path:path}")
    def serve_spa(full_path: str):
        return _FileResponse(_STATIC_DIR / "index.html")
